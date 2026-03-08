import io
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..tools.blob_storage_tool import upload_bytes_to_blob, generate_sas_url
from ..config import get_settings

settings = get_settings()


def generate_and_upload_excel(
    columns: list[str],
    rows: list[list],
    admin_user_id: int,
) -> dict:
    """
    Generate an Excel file from query results and upload to Blob Storage.
    Returns {download_url, file_name, blob_path}.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Asset Query Result"

    # Header row styling
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto column width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    excel_bytes = buffer.read()

    # Upload to Blob Storage
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    file_name = f"asset_query_{timestamp}.xlsx"
    blob_path = f"exports/{admin_user_id}/{file_name}"
    upload_bytes_to_blob(excel_bytes, blob_path, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    sas_url = generate_sas_url(blob_path, expiry_minutes=60)
    return {"download_url": sas_url, "file_name": file_name, "blob_path": blob_path}
